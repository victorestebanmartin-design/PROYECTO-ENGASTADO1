"""
Script temporal para analizar el archivo Excel
"""
import openpyxl
import sys

# Abrir el archivo Excel
try:
    workbook = openpyxl.load_workbook('data/LISTADO CABLEADO CORADIA ITALIA.xlsx', data_only=True)
    
    print("=" * 80)
    print("ANÁLISIS DEL ARCHIVO EXCEL")
    print("=" * 80)
    
    # Listar todas las hojas
    print(f"\n📋 HOJAS DISPONIBLES ({len(workbook.sheetnames)}):")
    for idx, sheet_name in enumerate(workbook.sheetnames, 1):
        print(f"  {idx}. {sheet_name}")
    
    # Analizar la primera hoja
    print("\n" + "=" * 80)
    sheet = workbook.active
    print(f"\n📊 ANÁLISIS DE LA HOJA ACTIVA: '{sheet.title}'")
    print("=" * 80)
    
    # Obtener dimensiones
    print(f"\n📏 Dimensiones:")
    print(f"   - Filas: {sheet.max_row}")
    print(f"   - Columnas: {sheet.max_column}")
    
    # Obtener nombres de columnas (primera fila)
    print(f"\n📌 COLUMNAS (Primera fila):")
    headers = []
    for col in range(1, min(sheet.max_column + 1, 30)):  # Máximo 30 columnas
        cell_value = sheet.cell(row=1, column=col).value
        if cell_value:
            headers.append(str(cell_value))
            print(f"   {col}. {cell_value}")
    
    # Mostrar primeras 5 filas de datos
    print(f"\n📄 PRIMERAS 5 FILAS DE DATOS:")
    print("-" * 80)
    
    for row in range(2, min(7, sheet.max_row + 1)):  # Filas 2-6
        print(f"\nFila {row}:")
        for col in range(1, min(len(headers) + 1, 15)):  # Primeras columnas
            cell_value = sheet.cell(row=row, column=col).value
            if col <= len(headers):
                print(f"   {headers[col-1]}: {cell_value}")
    
    print("\n" + "=" * 80)
    print("✅ Análisis completado")
    print("=" * 80)
    
    workbook.close()
    
except Exception as e:
    print(f"❌ Error al leer el archivo: {e}")
    sys.exit(1)
